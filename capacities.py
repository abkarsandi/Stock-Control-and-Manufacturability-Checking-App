import sys
import numpy as np
import pandas as pd
import openpyxl as opx
import json
import re
import os



def capacity(machine, diameter, tolerance, standard, length):

    with open(os.getcwd() + "\config.json") as f:
            config = json.load(f)
    
    limits_path = config["limits_file_path"]
    limits = opx.load_workbook(limits_path)

    ws = limits.active

    row = ws.iter_rows(2,max_row=2)

    for i in row:
        for cell in i:
            
            try:
                new = cell.value

            except:
                print("No value")

            if str(cell.value).lower() == str(machine).lower():
                machine_col = cell.column
                break
    
    try:
        print(isinstance(machine_col, int))

    except: 
        return "Makinaya dair bilgi bulunamadı."        

    if tolerance != "6g":
        machine_col +=1

    cell_coordinate = [4,machine_col]

    # Retrieve the value in the 4th row of the specified column
    diameter_cell = ws.cell(cell_coordinate[0],cell_coordinate[1])

    for i in ws.iter_rows(2):
        for cell in i:
            if standard == cell.value:
                std_row = cell.row

    try:
        if std_row is None:
            return "Standarda ait bilgi bulunamadı."
    except:
        print("Standart bilgisi bulunamadı.")    

    try:

        max = ws.cell(std_row, machine_col).value
        min = ws.cell(std_row+1, machine_col).value

        print("max: " + str(max) + ", min: " + str(min))

        max_num = int(max)
        min_num = int(min)
        len = int(length.strip())
        print(length)

        if diameter in diameter_cell.value and int(len) <= max_num and int(len) >=min_num:
            return ("Ürün bu makinede üretilebilir")

        else: return "Makine limitlerinin dışında"
    except: return "Veritabanında makine verisi mevcut değil"