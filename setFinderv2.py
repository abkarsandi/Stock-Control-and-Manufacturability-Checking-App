from PySide6.QtWidgets import QApplication, QMainWindow, QComboBox, QLabel, QPushButton, QVBoxLayout, QWidget, QScrollArea, QTextEdit, QLineEdit
import openpyxl as opx
import pandas as pd
import numpy as np
import json
import re

def setfinder(diameter, tolerance, thread, standard, length):
    

    with open("config.json") as f:
        config = json.load(f)

    list_path = config["list_file_path"]

    list = opx.load_workbook(list_path)
    df = pd.read_excel(list_path,dtype=str)

    ws = list.active

    dia_col = ws.iter_cols(2,max_col=2)
    standard_col = ws.iter_cols(4,max_col=4)

    dias = []
    stds = []
    die_sets = []

    machine = ""
    note = ""

    machine_and_notes = []

    for i in dia_col:
        for cell in i:
            if cell.value == diameter:
                dias.append(cell.row)

    for i in standard_col:
        for cell in i:
            if cell.row in dias:
                try: 
                    if re.search(standard, cell.value):
                        stds.append(cell.row)
                except:
                    print("Error at " + str(standard) + str(cell.value))

    print(stds)            
    for i in stds:
        try:
            if ws.cell(i,5).value == thread:
                try:
                    if re.search(tolerance, ws.cell(i,6).value):
                        machine = ws.cell(i,1).value
                        try:
                            if len(ws.cell(i,7).value) > 0:
                                note = ws.cell(i,7).value
                        except:
                            note = "Not bulunamadı"
                        
                        part_list = df.iloc[i-2][7:].values

                        try:
                            try:
                                exact = int(ws.cell(i,3).value)
                            except:
                                exactf = float(ws.cell(i,3).value)
                            if exact == length or exactf == length:
                                print("Length is exactly equal")
                                return [machine, note], part_list
                        except:
                            print("No specific length")
                            machine_and_note = [machine, note]
                            machine_and_notes.append(machine_and_note)
                            die_sets.append(part_list)
                except:
                    machine = ws.cell(i,1).value
                    try:
                        if len(ws.cell(i,7).value) > 0:
                            note = ws.cell(i,7).value
                    except:
                        note = "Not bulunamadı"
                    
                    

                    try:
                        value = ws.cell(i, 3).value
                        if isinstance(value, str):
                            value = float(value)  # Convert the string to float
                        if isinstance(value, (int, float)):
                            print(value)
                            if abs(value - float(length)) <= 0.0005:
                                part_list = (df.iloc[i-2][7:].values)
                                print("Length is exactly equal")
                                print(machine)
                                print(note)
                                machine_and_note = [machine, note]
                                machine_and_notes = []
                                die_sets = []
                                machine_and_notes.append(machine_and_note)
                                die_sets.append(part_list)
                                return machine_and_notes, die_sets
                        else:
                            print("No specific length")
                            part_list = df.iloc[i-2][7:].values
                            machine_and_note = [machine, note]
                            machine_and_notes.append(machine_and_note)
                            die_sets.append(part_list)
                    except Exception as e:
                        print("An error occurred:", str(e))

        except:
            print("Thread yok")
        
    return machine_and_notes, die_sets